import streamlit as st
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- KONFIGURASI HALAMAN STREAMLIT ---
st.set_page_config(page_title="Sistem Packing Walet Advance", layout="wide")

FILE_NAME = "data_packing_walet.xlsx"

DAFTAR_WALET = [
    "A", "A BC", "A JELEK", "A JUMBO", "A JUMBO @100", "A KUNING", "A NTR L", 
    "A NTR L SEMU", "A NTR T", "A SEMU", "A1.4", "A1.5", "A1.6", "A2.5", "A2.6", 
    "AP", "AP.5", "AP.6", "B", "B BC", "B JELEK", "B KUNING", "B NTR L", 
    "B NTR T", "B SEMU", "B1", "B2", "BP", "C BC", "C JELEK", "C KUNING", 
    "C NTR L", "C SEMU", "C1", "C2", "CP", "CT 1", "CT SEMU", "HCR @10", 
    "M1", "M1 H", "M2", "M2 H", "M2 MIX", "M3", "M3 H", "MK KCL", "MK KCL NTR", 
    "MK SEMU", "N1", "N1 BC", "N1 P", "N1+", "N2", "N2 BC", "OV", "OV 1", 
    "OV 2", "OV H", "OV JUMBO", "OV JUMBO @100", "OV JUMBO @50", "OV KCL", 
    "OV NTR L", "OV NTR T", "OV SEMU", "P1", "P1 @20", "P1 @25", "P1 GPG SDT", 
    "P1 GPG SDT H", "P1 H", "P1 MK", "P1 MK ABU", "P1 MK NTR", "P1 MK NTR SEMU", 
    "P1 MK SEMU", "P1 OV", "P1 OV NTR", "P1 SDT", "P1 SDT NTR", "P2", "P2 H", 
    "P2 NTR", "P3 @10", "S+ KUNING", "S+ SEMU", "SDT 1", "SDT 2", "SDT JUMBO", 
    "SDT JUMBO @100", "SDT JUMBO @50", "SDT KCL", "SDT KCL NTR", "SDT NTR L", 
    "SDT NTR T", "SDT SEMU", "TERI @10"
]

DAFTAR_CUSTOMER = ["AH-GN", "XM", "MR LIAO", "UMUM / TANPA NAMA"]
DAFTAR_KETERANGAN = ["STOK PACKING", "TIDAK PO", "OUTSPEK"]
DAFTAR_SURAT_JALAN = ["AH", "XM", "AH-GN", "MIX", "NTR", "TANPA SURAT JALAN"]

# --- FUNGSI EXCEL STYLING & GENERATION (SAMA SEPERTI VERSI LAMA) ---
def buat_border_dan_format(ws, min_r, max_r, min_c, max_c, font_data, thin_border):
    for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
        for cell in row:
            cell.font = font_data
            cell.border = thin_border
            if cell.column > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

def buat_stok_global_sheet(wb, df_transaksi):
    if "Stok_Global_All" in wb.sheetnames: del wb["Stok_Global_All"]
    ws = wb.create_sheet(title="Stok_Global_All")
    ws.views.sheetView[0].showGridLines = True
    ws.merge_cells("A1:D2")
    title_cell = ws["A1"]
    title_cell.value = "LAPORAN RANGKUMAN STOK GLOBAL PER KATEGORI KETERANGAN"
    title_cell.font = Font(name="Arial", size=13, bold=True, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"] = f"Update Data Terakhir: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"].font = Font(name="Arial", size=9, italic=True)
    
    current_row = 5
    fill_box_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_sub_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_total = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_outspek_box = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    font_box_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_sub = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)
    font_bold = Font(name="Arial", size=10, bold=True)
    thin_border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
                         top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))

    for ket in DAFTAR_KETERANGAN:
        df_ket = df_transaksi[df_transaksi["Keterangan"] == ket]
        items_aktif = df_ket["Nama Barang / Grade"].unique()
        data_kotak = []
        for item in DAFTAR_WALET:
            if item in items_aktif:
                df_item = df_ket[df_ket["Nama Barang / Grade"] == item]
                masuk = df_item[df_item["Status"] == "MASUK"]["Berat (Kg/Gram)"].sum()
                keluar = df_item[df_item["Status"] == "KELUAR"]["Berat (Kg/Gram)"].sum()
                sisa = masuk - keluar
                if masuk > 0 or keluar > 0 or sisa != 0:
                    data_kotak.append({"Grade": item, "Masuk": masuk, "Keluar": keluar, "Sisa": sisa})
        if not data_kotak: continue
            
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        hb = ws.cell(row=current_row, column=1, value=f"📌 KATEGORI: {ket}")
        hb.fill = fill_box_header; hb.font = font_box_header
        hb.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        current_row += 1
        
        headers_sub = ["Nama Barang / Grade", "Total Masuk (Kg)", "Total Keluar (Kg)", "Sisa Ready Stok"]
        for col_idx, h_text in enumerate(headers_sub, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=h_text)
            cell.fill = fill_sub_header if ket != "OUTSPEK" else fill_outspek_box
            cell.font = font_sub; cell.alignment = Alignment(horizontal="center"); cell.border = thin_border
        current_row += 1
        
        start_sum = current_row
        for row_data in data_kotak:
            ws.cell(row=current_row, column=1, value=row_data["Grade"]).alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=2, value=row_data["Masuk"]).alignment = Alignment(horizontal="right")
            ws.cell(row=current_row, column=3, value=row_data["Keluar"]).alignment = Alignment(horizontal="right")
            ws.cell(row=current_row, column=4, value=row_data["Sisa"]).alignment = Alignment(horizontal="right")
            buat_border_dan_format(ws, current_row, current_row, 1, 4, font_regular, thin_border)
            if ket == "OUTSPEK":
                for c in range(1, 5): ws.cell(row=current_row, column=c).fill = fill_outspek_box
            current_row += 1
        end_sum = current_row - 1
        
        t1 = ws.cell(row=current_row, column=1, value=f"TOTAL ALL {ket}")
        t2 = ws.cell(row=current_row, column=2, value=f"=SUM(B{start_sum}:B{end_sum})")
        t3 = ws.cell(row=current_row, column=3, value=f"=SUM(C{start_sum}:C{end_sum})")
        t4 = ws.cell(row=current_row, column=4, value=f"=SUM(D{start_sum}:D{end_sum})")
        t1.alignment = Alignment(horizontal="center")
        for t in [t1, t2, t3, t4]:
            t.font = font_bold; t.fill = fill_total if ket != "OUTSPEK" else fill_outspek_box; t.border = thin_border
            if t != t1: t.alignment = Alignment(horizontal="right"); t.number_format = '#,##0.00'
        current_row += 3
    for col in range(1, 5): ws.column_dimensions[get_column_letter(col)].width = 24

def buat_daily_report_sheet(wb, df_transaksi):
    if "Laporan_Harian_Daily" in wb.sheetnames: del wb["Laporan_Harian_Daily"]
    ws = wb.create_sheet(title="Laporan_Harian_Daily")
    ws.views.sheetView[0].showGridLines = True
    hari_ini = datetime.now().strftime("%Y-%m-%d")
    
    ws.merge_cells("A1:D2")
    title_cell = ws["A1"]; title_cell.value = f"LAPORAN TRANS-HARIAN DAILY PER KETERANGAN ({hari_ini})"
    title_cell.font = Font(name="Arial", size=13, bold=True, color="7F6000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    df_transaksi['Tanggal_Hari'] = df_transaksi['Tanggal'].astype(str).str.split(' ').str[0]
    df_hari_ini = df_transaksi[df_transaksi['Tanggal_Hari'] == hari_ini]
    
    current_row = 5
    fill_daily_header = PatternFill(start_color="7F6000", end_color="7F6000", fill_type="solid")
    fill_sub = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_outspek_box = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True); font_regular = Font(name="Arial", size=10)
    thin_border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
                         top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))

    for ket in DAFTAR_KETERANGAN:
        df_ket_day = df_hari_ini[df_hari_ini["Keterangan"] == ket]
        items_aktif_day = df_ket_day["Nama Barang / Grade"].unique()
        data_kotak_day = []
        for item in DAFTAR_WALET:
            if item in items_aktif_day:
                df_item_day = df_ket_day[df_ket_day["Nama Barang / Grade"] == item]
                masuk = df_item_day[df_item_day["Status"] == "MASUK"]["Berat (Kg/Gram)"].sum()
                keluar = df_item_day[df_item_day["Status"] == "KELUAR"]["Berat (Kg/Gram)"].sum()
                if masuk > 0 or keluar > 0: data_kotak_day.append({"Grade": item, "Masuk": masuk, "Keluar": keluar})
        if not data_kotak_day: continue
            
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        hd = ws.cell(row=current_row, column=1, value=f"📋 DAILY ACTIVITY: {ket}")
        hd.fill = fill_daily_header; hd.font = font_header; hd.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        current_row += 1
        
        headers_sub = ["Nama Barang / Grade", "Total Masuk Hari Ini", "Total Keluar Hari Ini"]
        for col_idx, text in enumerate(headers_sub, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=text)
            c.fill = fill_sub if ket != "OUTSPEK" else fill_outspek_box
            c.font = font_bold; c.alignment = Alignment(horizontal="center"); c.border = thin_border
        current_row += 1
        
        start_sum_d = current_row
        for row_data in data_kotak_day:
            ws.cell(row=current_row, column=1, value=row_data["Grade"]).alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=2, value=row_data["Masuk"]).alignment = Alignment(horizontal="right")
            ws.cell(row=current_row, column=3, value=row_data["Keluar"]).alignment = Alignment(horizontal="right")
            buat_border_dan_format(ws, current_row, current_row, 1, 3, font_regular, thin_border)
            if ket == "OUTSPEK":
                for c in range(1, 4): ws.cell(row=current_row, column=c).fill = fill_outspek_box
            current_row += 1
        end_sum_d = current_row - 1
        
        t1 = ws.cell(row=current_row, column=1, value=f"TOTAL AKTIVITAS {ket}")
        t2 = ws.cell(row=current_row, column=2, value=f"=SUM(B{start_sum_d}:B{end_sum_d})")
        t3 = ws.cell(row=current_row, column=3, value=f"=SUM(C{start_sum_d}:C{end_sum_d})")
        for t in [t1, t2, t3]:
            t.font = font_bold; t.fill = fill_sub if ket != "OUTSPEK" else fill_outspek_box; t.border = thin_border
            if t != t1: t.alignment = Alignment(horizontal="right"); t.number_format = '#,##0.00'
        current_row += 3
        
    if 'Tanggal_Hari' in df_transaksi.columns: df_transaksi = df_transaksi.drop(columns=['Tanggal_Hari'])
    for col in range(1, 4): ws.column_dimensions[get_column_letter(col)].width = 25

def buat_packing_list_sheet(wb, df_transaksi):
    if "Packing_List_Per_Customer" in wb.sheetnames: del wb["Packing_List_Per_Customer"]
    ws = wb.create_sheet(title="Packing_List_Per_Customer")
    ws.views.sheetView[0].showGridLines = True
    ws.merge_cells("A1:D2")
    ws["A1"] = "LAPORAN PACKING LIST - RINGKASAN STOK PER CUSTOMER"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    current_row = 5
    fill_cust_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_tbl_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_tbl_total = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    font_cust_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_tbl_header = Font(name="Arial", size=10, bold=True)
    font_bold = Font(name="Arial", size=10, bold=True); font_regular = Font(name="Arial", size=10)
    thin_border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
                               top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))

    for cust in DAFTAR_CUSTOMER:
        df_cust = df_transaksi[df_transaksi["Customer"] == cust]
        items_aktif = df_cust["Nama Barang / Grade"].unique()
        data_box = []
        for item in DAFTAR_WALET:
            if item in items_aktif:
                df_item = df_cust[df_cust["Nama Barang / Grade"] == item]
                masuk = df_item[df_item["Status"] == "MASUK"]["Berat (Kg/Gram)"].sum()
                keluar = df_item[df_item["Status"] == "KELUAR"]["Berat (Kg/Gram)"].sum()
                sisa = masuk - keluar
                if sisa != 0 or masuk > 0: data_box.append({"Grade": item, "Masuk": masuk, "Keluar": keluar, "Sisa": sisa})
        if not data_box: continue
            
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        h_cell = ws.cell(row=current_row, column=1, value=f"📦 CUSTOMER: {cust}")
        h_cell.fill = fill_cust_header; h_cell.font = font_cust_header; h_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        current_row += 1
        
        headers_sub = ["Grade Walet", "Total Masuk (Kg)", "Total Keluar (Kg)", "Sisa Stok (Ready)"]
        for col_idx, h_text in enumerate(headers_sub, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=h_text)
            cell.fill = fill_tbl_header; cell.font = font_tbl_header; cell.alignment = Alignment(horizontal="center"); cell.border = thin_border
        current_row += 1
        
        start_sum = current_row
        for row_data in data_box:
            ws.cell(row=current_row, column=1, value=row_data["Grade"]).alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=2, value=row_data["Masuk"]).alignment = Alignment(horizontal="right")
            ws.cell(row=current_row, column=3, value=row_data["Keluar"]).alignment = Alignment(horizontal="right")
            ws.cell(row=current_row, column=4, value=row_data["Sisa"]).alignment = Alignment(horizontal="right")
            buat_border_dan_format(ws, current_row, current_row, 1, 4, font_regular, thin_border)
            current_row += 1
        end_sum = current_row - 1
        
        t1 = ws.cell(row=current_row, column=1, value=f"TOTAL STOK {cust}")
        t2 = ws.cell(row=current_row, column=2, value=f"=SUM(B{start_sum}:B{end_sum})")
        t3 = ws.cell(row=current_row, column=3, value=f"=SUM(C{start_sum}:C{end_sum})")
        t4 = ws.cell(row=current_row, column=4, value=f"=SUM(D{start_sum}:D{end_sum})")
        for t in [t1, t2, t3, t4]:
            t.font = font_bold; t.fill = fill_tbl_total; t.border = thin_border
            if t != t1: t.alignment = Alignment(horizontal="right"); t.number_format = '#,##0.00'
        current_row += 3
    for col in range(1, 5): ws.column_dimensions[get_column_letter(col)].width = 22

def perindah_excel():
    if not os.path.exists(FILE_NAME): return
    try:
        wb = load_workbook(FILE_NAME)
        if "Data_Transaksi" in wb.sheetnames:
            ws = wb["Data_Transaksi"]
            ws.views.sheetView[0].showGridLines = True
            fill_h = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            font_h = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            thin_border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
                                 top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
            for cell in ws[1]:
                cell.fill = fill_h; cell.font = font_h; cell.alignment = Alignment(horizontal='center')
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border; cell.font = Font(name="Arial", size=10)
                    if isinstance(cell.value, (int, float)): cell.number_format = '#,##0.00'
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)
        wb.save(FILE_NAME)
    except: pass

def hitung_dan_simpan_excel(df_transaksi):
    kolom_standar = ["Tanggal", "Nama Barang / Grade", "Customer", "Surat Jalan", "Berat (Kg/Gram)", "Status", "Keterangan"]
    df_transaksi = df_transaksi[[c for c in kolom_standar if c in df_transaksi.columns]]
    try:
        with pd.ExcelWriter(FILE_NAME, engine='openpyxl') as writer:
            df_transaksi.to_excel(writer, sheet_name="Data_Transaksi", index=False)
            buat_stok_global_sheet(writer.book, df_transaksi)
            buat_daily_report_sheet(writer.book, df_transaksi)
            buat_packing_list_sheet(writer.book, df_transaksi)
        perindah_excel()
        return True
    except Exception as e:
        st.error(f"Gagal menyimpan ke Excel! Pastikan file tidak dibuka di program lain. Detail: {e}")
        return False

# --- LOGIKA AMBIL DATA LOKAL ---
if os.path.exists(FILE_NAME):
    try: df_global = pd.read_excel(FILE_NAME, sheet_name="Data_Transaksi")
    except: df_global = pd.DataFrame(columns=["Tanggal", "Nama Barang / Grade", "Customer", "Surat Jalan", "Berat (Kg/Gram)", "Status", "Keterangan"])
else:
    df_global = pd.DataFrame(columns=["Tanggal", "Nama Barang / Grade", "Customer", "Surat Jalan", "Berat (Kg/Gram)", "Status", "Keterangan"])

# --- ANTARMUKA WEB STREAMLIT ---
st.title("🦅 Sistem Pencatatan & Packing Walet Advance")
st.write("Aplikasi Manajemen Stok Web Server berbasis Python Streamlit.")

# Menggunakan Fitur Tab bawaan Streamlit (Otomatis Sangat Rapi)
tab1, tab2, tab3 = st.tabs(["📝 Form Input Utama", "🔄 Form Pindah Barang (Mutasi)", "🔎 Cari & Filter Log Log Data"])

# ==================== TAB 1: INPUT UTAMA ====================
with tab1:
    st.subheader("Form Entri Transaksi Masuk/Keluar")
    col1, col2 = st.columns(2)
    
    with col1:
        nama_barang = st.selectbox("Pilih Grade Walet:", DAFTAR_WALET, key="in_nama")
        customer = st.selectbox("Pilih Customer:", DAFTAR_CUSTOMER, key="in_cust")
        surat_jalan = st.selectbox("Pilih Jenis Surat Jalan:", DAFTAR_SURAT_JALAN, key="in_sj")
    
    with col2:
        keterangan = st.selectbox("Pilih Keterangan:", DAFTAR_KETERANGAN, key="in_ket")
        berat = st.number_input("Masukkan Berat (Kg/Gram):", min_value=0.0, step=0.01, format="%.2f", key="in_berat")
        status = st.radio("Pilih Status Barang:", ["MASUK", "KELUAR"], horizontal=True)

    if st.button("SIMPAN DATA TRANSAKSI", type="primary"):
        if berat <= 0:
            st.warning("Berat harus lebih besar dari 0!")
        else:
            tanggal_jam = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            df_baru = pd.DataFrame([{"Tanggal": tanggal_jam, "Nama Barang / Grade": nama_barang, "Customer": customer,
                                     "Surat Jalan": surat_jalan, "Berat (Kg/Gram)": berat, "Status": status, "Keterangan": keterangan}])
            df_global = pd.concat([df_global, df_baru], ignore_index=True)
            if hitung_dan_simpan_excel(df_global):
                st.success(f"Sukses tersimpan: {nama_barang} sebanyak {berat} Kg ke Excel!")
                st.rerun()

# ==================== TAB 2: PINDAH BARANG ====================
with tab2:
    st.subheader("Form Mutasi Stok Antar Customer")
    col1, col2 = st.columns(2)
    
    with col1:
        p_nama = st.selectbox("Pilih Grade Walet yang Dipindah:", DAFTAR_WALET, key="p_nama")
        p_asal = st.selectbox("Pilih Customer ASAL (Stok Berkurang):", DAFTAR_CUSTOMER, key="p_asal")
        p_tujuan = st.selectbox("Pilih Customer TUJUAN (Stok Bertambah):", DAFTAR_CUSTOMER, key="p_tujuan")
        
    with col2:
        p_sj = st.selectbox("Pilih Jenis Surat Jalan Mutasi:", DAFTAR_SURAT_JALAN, key="p_sj")
        p_ket = st.selectbox("Pilih Keterangan Mutasi:", DAFTAR_KETERANGAN, key="p_ket")
        p_berat = st.number_input("Masukkan Berat Mutasi (Kg/Gram):", min_value=0.0, step=0.01, format="%.2f", key="p_berat")

    if st.button("PROSES PINDAH BARANG", type="primary"):
        if p_asal == p_tujuan:
            st.error("Customer Asal dan Tujuan tidak boleh sama!")
        elif p_berat <= 0:
            st.warning("Berat mutasi harus lebih besar dari 0!")
        else:
            tanggal_jam = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            data_keluar = {"Tanggal": tanggal_jam, "Nama Barang / Grade": p_nama, "Customer": p_asal,
                           "Surat Jalan": p_sj, "Berat (Kg/Gram)": p_berat, "Status": "KELUAR", "Keterangan": p_ket}
            data_masuk = {"Tanggal": tanggal_jam, "Nama Barang / Grade": p_nama, "Customer": p_tujuan,
                          "Surat Jalan": p_sj, "Berat (Kg/Gram)": p_berat, "Status": "MASUK", "Keterangan": p_ket}
            
            df_global = pd.concat([df_global, pd.DataFrame([data_keluar, data_masuk])], ignore_index=True)
            if hitung_dan_simpan_excel(df_global):
                st.success(f"Berhasil mutasi {p_berat} Kg {p_nama} dari {p_asal} ke {p_tujuan}!")
                st.rerun()

# ==================== TAB 3: CARI & FILTER LOG DATA ====================
with tab3:
    st.subheader("Filter dan Riwayat Seluruh Transaksi")
    
    f_col1, f_col2, f_col3 = st.columns(3)
    with f_col1:
        f_grade = st.text_input("Ketik Nama Grade (Cari Sebagian Nama):").upper().strip()
    with f_col2:
        f_cust = st.selectbox("Filter Customer:", ["SEMUA CUSTOMER"] + DAFTAR_CUSTOMER)
    with f_col3:
        f_status = st.selectbox("Filter Status:", ["SEMUA STATUS", "MASUK", "KELUAR"])
        
    df_filtered = df_global.copy()
    if f_grade:
        df_filtered = df_filtered[df_filtered["Nama Barang / Grade"].astype(str).str.contains(f_grade, na=False)]
    if f_cust != "SEMUA CUSTOMER":
        df_filtered = df_filtered[df_filtered["Customer"] == f_cust]
    if f_status != "SEMUA STATUS":
        df_filtered = df_filtered[df_filtered["Status"] == f_status]
        
    df_tampil = df_filtered.iloc[::-1].reset_index() # Urutan terbaru di atas
    
    # Menampilkan Tabel di Halaman Web
    st.dataframe(df_tampil[["Tanggal", "Nama Barang / Grade", "Customer", "Surat Jalan", "Berat (Kg/Gram)", "Status", "Keterangan"]], use_container_width=True)
    
    # Fitur Hapus Baris Terpilih Terintegrasi Aman
    st.write("---")
    st.subheader("🗑 Hapus Transaksi")
    if not df_tampil.empty:
        opsi_hapus = [f"Index {row['index']} | {row['Tanggal']} | {row['Nama Barang / Grade']} | {row['Berat (Kg/Gram)']} Kg ({row['Status']})" for _, row in df_tampil.iterrows()]
        pilihan_hapus = st.selectbox("Pilih baris transaksi yang ingin dihapus permanent:", ["-- PILIH DATA --"] + opsi_hapus)
        
        if st.button("HAPUS DATA TERPILIH", type="secondary"):
            if pilihan_hapus != "-- PILIH DATA --":
                id_baris_asli = int(pilihan_hapus.split(" ")[1])
                df_global = df_global.drop(index=id_baris_asli).reset_index(drop=True)
                if hitung_dan_simpan_excel(df_global):
                    st.success("Data berhasil dihapus dari sistem database!")
                    st.rerun()
            else:
                st.warning("Silakan pilih baris transaksi terlebih dahulu!")

# --- FOOTER MINI TABEL (5 TRANSAKSI TERAKHIR) ---
st.write("---")
st.markdown("### ⏱ 5 Transaksi Terakhir (Live Activity Log)")
st.table(df_global.tail(5).iloc[::-1][["Tanggal", "Nama Barang / Grade", "Customer", "Berat (Kg/Gram)", "Status"]])
